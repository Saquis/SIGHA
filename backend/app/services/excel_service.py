# German Del Rio
# Desarrollador Version 1
# SIGHA - Sistema de Gestión de Horarios y Asignación
# services/excel_service.py: Generación de reportes Excel institucionales

import io
from openpyxl import Workbook
from sqlalchemy.orm import Session
from app.models import Horario

def generar_reporte_excel(db: Session):
    wb = Workbook()
    wb.remove(wb.active) # Remover la hoja por defecto

    # 1. Crear las 5 hojas exactas solicitadas por el ITQ
    ws_mod1 = wb.create_sheet(title="Modulo1")
    ws_mod2 = wb.create_sheet(title="Modulo2")
    ws_mod3 = wb.create_sheet(title="Modulo3")
    ws_carga_mod = wb.create_sheet(title="Carga horaria por módulos")
    ws_carga_total = wb.create_sheet(title="Carga horaria total")

    # 2. Configurar cabeceras para las hojas de Módulos
    cabeceras_mod = ["Docente", "Asignatura", "Carrera", "Paralelo", "Jornada", "Día", "Hora Inicio", "Hora Fin"]
    for ws in [ws_mod1, ws_mod2, ws_mod3]:
        ws.append(cabeceras_mod)

    # 3. Configurar cabeceras para las hojas de Carga Horaria
    ws_carga_mod.append(["Docente", "Módulo 1 (Horas)", "Módulo 2 (Horas)", "Módulo 3 (Horas)"])
    ws_carga_total.append(["Docente", "Tipo Contrato", "Total Horas", "Estado de Regla (272-380h)"])

    # Obtener todos los horarios
    horarios = db.query(Horario).all()
    
    # Diccionarios para agrupar las horas: {docente_nombre: {1: 0, 2: 0, 3: 0}}
    carga_docentes = {}
    info_docentes = {}

    # 4. Llenar las hojas de Módulos y sumar horas
    for h in horarios:
        docente_nombre = h.docente.usuario.nombre
        modulo_num = h.modulo.numero
        
        if docente_nombre not in carga_docentes:
            carga_docentes[docente_nombre] = {1: 0, 2: 0, 3: 0}
            info_docentes[docente_nombre] = h.docente.tipo

        fila_horario = [
            docente_nombre, 
            h.asignatura.nombre, 
            h.asignatura.carrera.nombre, 
            h.paralelo, 
            h.jornada, 
            h.dia, 
            h.hora_inicio.strftime("%H:%M"), 
            h.hora_fin.strftime("%H:%M")
        ]
        
        # Distribuir en la hoja correcta
        if modulo_num == 1:
            ws_mod1.append(fila_horario)
        elif modulo_num == 2:
            ws_mod2.append(fila_horario)
        elif modulo_num == 3:
            ws_mod3.append(fila_horario)

        # Sumar las horas semanales de la materia asignada
        carga_docentes[docente_nombre][modulo_num] += h.asignatura.horas_semanales

    # 5. Llenar las hojas de reportes consolidados
    for docente, cargas in carga_docentes.items():
        # Llenar: Carga horaria por módulos
        ws_carga_mod.append([docente, cargas[1], cargas[2], cargas[3]])
        
        # Llenar: Carga horaria total
        total_horas = cargas[1] + cargas[2] + cargas[3]
        tipo = info_docentes[docente]
        
        # Validación visual para el excel sobre la regla del ITQ
        estado = "CUMPLE"
        if tipo == "tiempo_completo":
            if total_horas < 272:
                estado = "ALERTA: Faltan horas"
            elif total_horas > 380:
                estado = "ALERTA: Excede límite"
                
        ws_carga_total.append([docente, tipo, total_horas, estado])

    # 6. Guardar en memoria (BytesIO) para enviarlo por la API sin crear archivos basura en el servidor
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream